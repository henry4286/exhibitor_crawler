"""
Excel导出模块

负责将爬取的数据保存到Excel文件
"""

import os
import threading
import time
from typing import Optional

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class ExcelExporter:
    """
    线程安全的Excel导出器
    
    负责将公司信息保存到Excel文件，支持多线程并发写入。
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录路径，默认为脚本上一级目录的ExhibitorList文件夹
        """
        if output_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(script_dir, "..", "ExhibitorList")
        
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 文件锁字典，每个文件一个锁
        self._file_locks = {}
        self._lock_for_locks = threading.Lock()
    
    def _get_file_lock(self, file_path: str) -> threading.Lock:
        """获取指定文件的锁"""
        with self._lock_for_locks:
            if file_path not in self._file_locks:
                self._file_locks[file_path] = threading.Lock()
            return self._file_locks[file_path]
    
    def save(self, company_list: list[dict], exhibition_code: str, headers: list[str]) -> bool:
        """
        保存公司列表到Excel文件（线程安全）
        
        Args:
            company_list: 公司信息列表
            exhibition_code: 展会代码（用作文件名）
            headers: 表头字段列表
        
        Returns:
            是否保存成功
        """
        if not company_list:
            return True
        
        file_path = os.path.join(self.output_dir, f"{exhibition_code}.xlsx")
        file_lock = self._get_file_lock(file_path)
        
        # 使用文件锁保护写入操作
        with file_lock:
            # 重试机制，防止文件被占用
            max_retries = 3
            retry_delay = 0.5
            
            for attempt in range(max_retries):
                try:
                    # 加载或创建工作簿
                    if os.path.exists(file_path):
                        workbook = load_workbook(file_path)
                        worksheet = workbook.active
                    else:
                        workbook = Workbook()
                        worksheet = workbook.active
                        if worksheet is not None:
                            worksheet.append(headers)
                    
                    # 写入数据行
                    if worksheet is not None:
                        for company in company_list:
                            # 直接获取数据，不添加前缀空格，避免Excel格式错误
                            row_data = []
                            for header in headers:
                                value = company.get(header, '')
                                # 清理控制字符，避免Excel保存错误
                                if isinstance(value, str):
                                    # 移除控制字符（ASCII 0-31，除了制表符、换行符、回车符）
                                    cleaned_value = ''.join(char for char in value if ord(char) >= 32 or char in '\t\n\r')
                                    row_data.append(cleaned_value)
                                else:
                                    row_data.append(value)
                            worksheet.append(row_data)
                    
                    workbook.save(file_path)
                    return True
                    
                except PermissionError:
                    if attempt < max_retries - 1:
                        print(f"文件 {file_path} 正在被占用，{retry_delay}秒后重试...", flush=True)
                        time.sleep(retry_delay)
                    else:
                        print(f"保存文件 {file_path} 失败，已超过最大重试次数", flush=True)
                        return False
                except Exception as e:
                    print(f"保存文件 {file_path} 时发生错误: {e}", flush=True)
                    return False
            
            return False
    
    def get_file_path(self, exhibition_code: str) -> str:
        """
        获取输出文件路径
        
        Args:
            exhibition_code: 展会代码
        
        Returns:
            完整的文件路径
        """
        return os.path.join(self.output_dir, f"{exhibition_code}.xlsx")
